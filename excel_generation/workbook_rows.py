"""Load existing Excel export rows for incremental updates."""

from pathlib import Path
from typing import Any

from configuration import OUTPUT_DIR
from openpyxl import load_workbook

DEFAULT_WORKBOOK_FILENAME = "workshop_catalog.xlsx"


def load_existing_workbook_rows(
    output_filename: str = DEFAULT_WORKBOOK_FILENAME,
) -> list[dict[str, Any]]:
    """Read rows from the current generated workbook if it exists."""
    workbook_path = Path(OUTPUT_DIR) / output_filename
    if not workbook_path.exists():
        return []

    workbook = load_workbook(
        filename=workbook_path,
        read_only=True,
        data_only=True,
    )
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration:
        workbook.close()
        return []

    loaded_rows: list[dict[str, Any]] = []
    for row_values in rows:
        row = {
            headers[index]: value
            for index, value in enumerate(row_values)
            if index < len(headers) and headers[index]
        }
        if any(value not in (None, "") for value in row.values()):
            loaded_rows.append(row)

    workbook.close()
    return loaded_rows
