import json
import logging
import os

from configuration import (
    DEFAULT_DATA_ROW_HEIGHT,
    EXCEL_CELL_CHARACTER_LIMIT,
    MINIMUM_COLUMN_WIDTH,
    OUTPUT_DIR,
    SLIDE_BREAK,
)
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def sanitize_excel_value(
    value, *, field_name: str | None = None, object_name: str | None = None
):
    if not isinstance(value, str):
        return value
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    if len(cleaned) > EXCEL_CELL_CHARACTER_LIMIT:
        label = object_name or "unknown object"
        column = field_name or "unknown field"
        logger.warning(
            (
                "Truncated text for '%s' in column '%s': %s chars exceeds Excel "
                "cell limit of %s chars."
            ),
            label,
            column,
            len(cleaned),
            EXCEL_CELL_CHARACTER_LIMIT,
        )
    return cleaned[:EXCEL_CELL_CHARACTER_LIMIT]


def serialize_object_for_excel(obj: dict) -> dict:
    serialized = {}
    object_name = str(obj.get("name", obj.get("id", "unknown object")))

    for key, value in obj.items():
        if key == "slide_texts":
            serialized[key] = sanitize_excel_value(
                SLIDE_BREAK.join(map(str, value)),
                field_name=key,
                object_name=object_name,
            )
        elif isinstance(value, list):
            serialized[key] = sanitize_excel_value(
                ", ".join(map(str, value)),
                field_name=key,
                object_name=object_name,
            )
        elif isinstance(value, (dict, tuple, set)):
            serialized[key] = sanitize_excel_value(
                json.dumps(value),
                field_name=key,
                object_name=object_name,
            )
        else:
            serialized[key] = sanitize_excel_value(
                value,
                field_name=key,
                object_name=object_name,
            )

    return serialized


def write_objects_to_excel(
    objects: list[dict],
    output_filename: str = "workshop_catalog.xlsx",
    headers: list[str] | None = None,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Presentations"

    output_path = os.path.join(OUTPUT_DIR, output_filename)

    if headers is None:
        headers = list(objects[0].keys()) if objects else []
    worksheet.append(headers)

    if not objects:
        workbook.save(output_path)
        return

    excel_objects = [serialize_object_for_excel(obj) for obj in objects]

    for column_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = max(
            len(str(header)) + 2, MINIMUM_COLUMN_WIDTH
        )

    for obj in excel_objects:
        worksheet.append([obj.get(header) for header in headers])

    for row in worksheet.iter_rows(min_row=2):
        worksheet.row_dimensions[row[0].row].height = DEFAULT_DATA_ROW_HEIGHT
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    workbook.save(output_path)
