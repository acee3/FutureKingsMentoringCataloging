"""Read uploaded workbooks and prepare rows for semantic indexing."""

from io import BytesIO

from openpyxl import load_workbook

from vector_search_app.types import IndexedWorkbookRow


def _normalize_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _choose_title(metadata: dict[str, str], fallback: str) -> str:
    for key in ("title", "name", "presentation_title", "file_name"):
        value = metadata.get(key, "").strip()
        if value:
            return value
    return fallback


def _build_searchable_text(metadata: dict[str, str]) -> str:
    preferred_fields = [
        "name",
        "theme*",
        "subtheme*",
        "description*",
        "audience*",
        "slide_texts",
        "presentation_path",
    ]
    ordered_keys: list[str] = []
    seen_keys: set[str] = set()
    for key in preferred_fields + list(metadata):
        if key in seen_keys:
            continue
        seen_keys.add(key)
        ordered_keys.append(key)

    parts: list[str] = []
    for key in ordered_keys:
        value = metadata.get(key, "").strip()
        if value:
            parts.append(f"{key}: {value}")
    return "\n".join(parts)


def load_index_rows(
    workbook_bytes: bytes,
    workbook_name: str,
    worksheet_name: str | None = None,
) -> list[IndexedWorkbookRow]:
    workbook = load_workbook(
        filename=BytesIO(workbook_bytes),
        read_only=True,
        data_only=True,
    )
    sheets = [workbook[worksheet_name]] if worksheet_name else list(workbook.worksheets)
    index_rows: list[IndexedWorkbookRow] = []

    for sheet in sheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [_normalize_cell_value(value) for value in rows[0]]
        for row_number, row_values in enumerate(rows[1:], start=2):
            metadata = {
                headers[index]: _normalize_cell_value(value)
                for index, value in enumerate(row_values)
                if index < len(headers) and headers[index]
            }
            if not any(metadata.values()):
                continue

            source_id = metadata.get("id") or f"{sheet.title}:{row_number}"
            title = _choose_title(metadata, source_id)
            index_rows.append(
                IndexedWorkbookRow(
                    source_id=source_id,
                    title=title,
                    workbook_name=workbook_name,
                    sheet_name=sheet.title,
                    row_number=row_number,
                    metadata=metadata,
                    searchable_text=_build_searchable_text(metadata),
                )
            )

    return index_rows
